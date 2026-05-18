"""
xlsx_writer.py — XLSX output orchestrator.
Creates multi-sheet report with matplotlib PNG charts embedded.
"""

import pandas as pd
import numpy as np
import json
import io
import os
from pathlib import Path
from datetime import datetime

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

BASE_DIR = Path(__file__).parent.parent.parent
CONFIG_PATH = BASE_DIR / "config.json"

# ─── DESIGN SYSTEM ───────────────────────────────────────────────────────────
COLORS = {
    'navy':    '1B2A4A',
    'green':   '2D7D46',
    'cyan':    '0097A7',
    'orange':  'E65100',
    'purple':  '6A1B9A',
    'skyblue': '0277BD',
    'pink':    'AD1457',
    'yellow':  'F9A825',
    'burgundy':'6D1C1C',
    'darkgreen':'1B5E20',
    'gray':    '424242',
    'muted':   '757575',
    'white':   'FFFFFF',
    'lightgray':'F5F5F5',
    'pass_green': 'C8E6C9',
    'warn_yellow':'FFF9C4',
    'fail_red':   'FFCDD2',
    'header_text':'FFFFFF',
}

STATUS_FILL = {
    'PASS': PatternFill('solid', fgColor=COLORS['pass_green']),
    'WARN': PatternFill('solid', fgColor=COLORS['warn_yellow']),
    'FAIL': PatternFill('solid', fgColor=COLORS['fail_red']),
    'NO_PARENT': PatternFill('solid', fgColor='E0E0E0'),
}

TAB_COLORS = {
    'DATA_CLEAN': COLORS['navy'],
    'QC_DUPLICATES': COLORS['green'],
    'QC_BLANKS': COLORS['cyan'],
    'QC_STANDARDS': COLORS['orange'],
    'RESSOURCES': COLORS['purple'],
    'COMP_RUNS': COLORS['skyblue'],
    'COMP_METHODES': COLORS['pink'],
    'COMP_LABOS': COLORS['yellow'],
    'COMPOSITE': COLORS['burgundy'],
    'COMP_CAMPAGNES': COLORS['darkgreen'],
    'HISTORIQUE': COLORS['gray'],
}


def _make_header_fill(hex_color: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_color)


def _header_font(bold=True) -> Font:
    return Font(name='Calibri', bold=bold, color=COLORS['white'], size=11)


def _data_font(bold=False) -> Font:
    return Font(name='Calibri', bold=bold, size=10)


def _thin_border():
    side = Side(style='thin', color='BDBDBD')
    return Border(left=side, right=side, top=side, bottom=side)


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _write_header_row(ws, row: int, headers: list, hex_color: str):
    fill = _make_header_fill(hex_color)
    font = _header_font()
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border()


def _status_cell(ws, row: int, col: int, status: str):
    cell = ws.cell(row=row, column=col, value=status)
    cell.fill = STATUS_FILL.get(status, PatternFill())
    cell.font = Font(name='Calibri', bold=True, size=10)
    cell.alignment = Alignment(horizontal='center')
    cell.border = _thin_border()


def _img_to_xl(fig) -> XLImage:
    """Convert matplotlib figure to openpyxl Image."""
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    buf.seek(0)
    img = XLImage(buf)
    plt.close(fig)
    return img


# ─── CHART GENERATORS ────────────────────────────────────────────────────────

def _chart_rpd_bar(pairs: list, element: str, thresholds: dict) -> plt.Figure:
    """Bar chart of RPD per duplicate pair for one element."""
    labels, rpds, colors = [], [], []
    for p in pairs:
        rv = p.get('rpd_values', {}).get(element)
        if rv:
            labels.append(p['dup_id'])
            rpds.append(rv['rpd'])
            st = rv['status']
            colors.append('#4CAF50' if st == 'PASS' else '#FF9800' if st == 'WARN' else '#F44336')

    if not labels:
        return None

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.bar(range(len(labels)), rpds, color=colors, edgecolor='white', linewidth=0.5)
    ax.axhline(thresholds['warn'], color='#FF9800', linestyle='--', linewidth=1.5,
               label=f'Warning {thresholds["warn"]}%')
    ax.axhline(thresholds['fail'], color='#F44336', linestyle='--', linewidth=1.5,
               label=f'Fail {thresholds["fail"]}%')
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
    ax.set_ylabel('RPD (%)', fontsize=10)
    ax.set_title(f'Duplicate RPD — {element}', fontsize=12, fontweight='bold')
    ax.legend(fontsize=8)
    ax.set_ylim(bottom=0)
    fig.patch.set_facecolor('white')
    ax.set_facecolor('#FAFAFA')
    fig.tight_layout()
    return fig


def _chart_scatter_regression(chart_data: dict) -> plt.Figure:
    """Scatter plot original vs duplicate with regression line."""
    x = chart_data['x']
    y = chart_data['y']
    el = chart_data['element']
    slope = chart_data['slope']
    intercept = chart_data['intercept']
    r2 = chart_data['r2']

    fig, ax = plt.subplots(figsize=(6, 6))
    ax.scatter(x, y, color='#1B2A4A', s=60, alpha=0.8, zorder=5)

    # 1:1 line
    mn = min(min(x), min(y))
    mx = max(max(x), max(y))
    ax.plot([mn, mx], [mn, mx], 'k--', linewidth=1, alpha=0.4, label='1:1 line')

    # Regression
    x_line = np.linspace(mn, mx, 100)
    y_line = slope * x_line + intercept
    ax.plot(x_line, y_line, color='#E65100', linewidth=2,
            label=f'y={slope:.3f}x+{intercept:.2f}\nR²={r2:.4f}')

    ax.set_xlabel('Original (ppm)', fontsize=10)
    ax.set_ylabel('Duplicate (ppm)', fontsize=10)
    ax.set_title(f'{el} — Original vs Duplicate', fontsize=11, fontweight='bold')
    ax.legend(fontsize=8)
    ax.set_facecolor('#FAFAFA')
    fig.patch.set_facecolor('white')
    fig.tight_layout()
    return fig


def _chart_shewhart(shewhart_data: dict, crm_id: str, element: str,
                    is_recovery=False) -> plt.Figure:
    """Shewhart control chart."""
    sd = shewhart_data
    values = sd['recoveries'] if is_recovery else sd.get('values', [])
    if not values:
        return None

    fig, ax = plt.subplots(figsize=(10, 4))
    x = range(1, len(values) + 1)

    # Control lines
    target = sd['target']
    ucl = sd['ucl']
    uwl = sd['uwl']
    lcl = sd['lcl']
    lwl = sd['lwl']

    ax.axhline(target, color='#2D7D46', linewidth=2, label='Target', zorder=2)
    ax.axhline(ucl, color='#F44336', linewidth=1.5, linestyle='--', label='UCL/LCL', zorder=2)
    ax.axhline(lcl, color='#F44336', linewidth=1.5, linestyle='--', zorder=2)
    ax.axhline(uwl, color='#FF9800', linewidth=1, linestyle=':', label='UWL/LWL', zorder=2)
    ax.axhline(lwl, color='#FF9800', linewidth=1, linestyle=':', zorder=2)

    # Fill zones
    ax.fill_between(range(0, len(values)+2), lcl, ucl, alpha=0.05, color='#F44336')
    ax.fill_between(range(0, len(values)+2), lwl, uwl, alpha=0.08, color='#FF9800')
    ax.fill_between(range(0, len(values)+2), lwl, uwl, alpha=0.08, color='#2D7D46')

    # Points colored by status
    point_colors = []
    for v in values:
        if v < lcl or v > ucl:
            point_colors.append('#F44336')
        elif v < lwl or v > uwl:
            point_colors.append('#FF9800')
        else:
            point_colors.append('#2D7D46')

    ax.plot(x, values, '-o', color='#1B2A4A', linewidth=1.5, zorder=5, alpha=0.6)
    ax.scatter(x, values, c=point_colors, s=80, zorder=6, edgecolors='white', linewidth=0.5)

    ylabel = '% Recovery' if is_recovery else 'Value (ppm)'
    ax.set_ylabel(ylabel, fontsize=10)
    ax.set_xlabel('Measurement #', fontsize=10)
    ax.set_title(f'{crm_id} — {element} Shewhart Chart', fontsize=11, fontweight='bold')
    ax.legend(fontsize=8, loc='upper right')
    ax.set_facecolor('#FAFAFA')
    fig.patch.set_facecolor('white')
    fig.tight_layout()
    return fig


def _chart_element_histogram(df: pd.DataFrame, element: str) -> plt.Figure:
    """Histogram of element distribution in ORIG samples."""
    col = f'{element}_ppm'
    if col not in df.columns:
        return None
    orig = df[df['SampleType'] == 'ORIG'][col].dropna()
    if len(orig) < 3:
        return None

    fig, ax = plt.subplots(figsize=(8, 4))
    ax.hist(orig, bins=30, color='#1B2A4A', edgecolor='white', linewidth=0.5, alpha=0.85)
    ax.axvline(orig.median(), color='#E65100', linewidth=2, linestyle='--',
               label=f'Median: {orig.median():.2f}')
    ax.axvline(orig.mean(), color='#FF9800', linewidth=2, linestyle=':',
               label=f'Mean: {orig.mean():.2f}')
    ax.set_xlabel(f'{element} (ppm)', fontsize=10)
    ax.set_ylabel('Count', fontsize=10)
    ax.set_title(f'{element} Distribution — ORIG Samples (n={len(orig)})', fontsize=11, fontweight='bold')
    ax.legend(fontsize=8)
    ax.set_facecolor('#FAFAFA')
    fig.patch.set_facecolor('white')
    fig.tight_layout()
    return fig


# ─── SHEET WRITERS ───────────────────────────────────────────────────────────

def _write_data_clean(wb: Workbook, df: pd.DataFrame, cfg: dict):
    ws = wb.create_sheet('DATA_CLEAN')
    ws.sheet_properties.tabColor = TAB_COLORS['DATA_CLEAN']

    elem_prio = cfg.get('elements_prioritaires', ['Ta', 'Nb', 'Ti', 'Rb'])
    elem_active = cfg.get('elements_actifs', [])

    # Build column list: metadata first, then priority elements, then others
    meta_cols = ['SampleID', 'HoleID', 'SiteID', 'ActivityType', 'From_m', 'To_m',
                 'SampleType', 'QAQCType', 'IDParent', 'IDBlk', 'IDStd',
                 'Mass_Bulk_kg', 'Mass_HMC_g', 'Method', 'Instrument']
    meta_cols = [c for c in meta_cols if c in df.columns]

    elem_cols_prio = [f'{e}_ppm' for e in elem_prio if f'{e}_ppm' in df.columns]
    elem_cols_other = [c for c in df.columns if c.endswith('_ppm')
                       and c not in elem_cols_prio
                       and not any(x in c for x in ['LOD', 'outlier'])]
    flag_cols = [c for c in df.columns if 'LOD_flag' in c or 'outlier_flag' in c]

    all_cols = meta_cols + elem_cols_prio + elem_cols_other
    headers = [c.replace('_ppm', '') if c.endswith('_ppm') else c for c in all_cols]

    # Title row
    ws.row_dimensions[1].height = 30
    title_cell = ws.cell(row=1, column=1, value='GeoQC Pro — DATA CLEAN')
    title_cell.font = Font(name='Calibri', bold=True, size=14, color=COLORS['navy'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(headers), 10))

    _write_header_row(ws, 2, headers, COLORS['navy'])
    ws.row_dimensions[2].height = 22

    # Freeze top 2 rows + first column
    ws.freeze_panes = ws.cell(row=3, column=2)

    # Data rows
    orig_fill = PatternFill('solid', fgColor='F8F9FA')
    qaqc_fill = PatternFill('solid', fgColor='E8F5E9')

    for ri, (_, row) in enumerate(df[all_cols].iterrows(), start=3):
        is_qaqc = str(row.get('SampleType', 'ORIG')) == 'QAQC'
        row_fill = qaqc_fill if is_qaqc else (orig_fill if ri % 2 == 0 else None)

        for ci, col in enumerate(all_cols, 1):
            val = row.get(col)
            if pd.isna(val) or val is None:
                val = ''
            cell = ws.cell(row=ri, column=ci, value=val if not isinstance(val, float) or not np.isnan(val) else '')
            if row_fill:
                cell.fill = row_fill
            cell.font = _data_font(bold=is_qaqc and ci == 1)
            cell.border = _thin_border()

            # Color status-based cells
            if col == 'SampleType':
                pass
            # Color numeric cells for priority elements by value (heat)
            if col in elem_cols_prio and isinstance(val, (int, float)) and val > 0:
                pass  # Could add conditional color scale here

    # Column widths
    col_widths = {'SampleID': 14, 'HoleID': 12, 'ActivityType': 12,
                  'From_m': 8, 'To_m': 8, 'SampleType': 10, 'QAQCType': 10,
                  'Mass_Bulk_kg': 12, 'Method': 14, 'Instrument': 18}
    for ci, col in enumerate(all_cols, 1):
        w = col_widths.get(col, 10 if col.endswith('_ppm') else 14)
        _set_col_width(ws, ci, w)

    # Add histograms for priority elements below data
    chart_row = len(df) + 5
    chart_col = 1
    for el in elem_prio[:4]:
        fig = _chart_element_histogram(df, el)
        if fig:
            img = _img_to_xl(fig)
            img.width = 480
            img.height = 240
            ws.add_image(img, f'{get_column_letter(chart_col)}{chart_row}')
            chart_col += 8


def _write_qc_duplicates(wb: Workbook, dup_results: dict, cfg: dict):
    ws = wb.create_sheet('QC_DUPLICATES')
    ws.sheet_properties.tabColor = TAB_COLORS['QC_DUPLICATES']

    pairs = dup_results.get('pairs', [])
    elem_prio = cfg.get('elements_prioritaires', ['Ta', 'Nb', 'Ti', 'Rb'])
    thresholds = dup_results.get('thresholds', {'warn': 10, 'fail': 20})

    # Title
    ws.cell(row=1, column=1, value='QC DUPLICATES — RPD Analysis').font = Font(
        name='Calibri', bold=True, size=14, color=COLORS['green'])

    # Summary row
    ws.cell(row=2, column=1, value=f"Pairs: {dup_results.get('n_pairs',0)} | "
            f"PASS: {dup_results.get('n_pass',0)} | "
            f"WARN: {dup_results.get('n_warn',0)} | "
            f"FAIL: {dup_results.get('n_fail',0)}")

    # Header: DUP_ID, PARENT_ID, STATUS, then elements ×3 (orig/dup/rpd)
    headers = ['DUP_ID', 'PARENT_ID', 'OVERALL_STATUS']
    for el in elem_prio:
        headers += [f'{el}_ORIG', f'{el}_DUP', f'{el}_RPD%', f'{el}_STATUS']

    _write_header_row(ws, 4, headers, COLORS['green'])

    for ri, pair in enumerate(pairs, start=5):
        ws.cell(row=ri, column=1, value=pair.get('dup_id'))
        ws.cell(row=ri, column=2, value=pair.get('parent_id'))
        _status_cell(ws, ri, 3, pair.get('overall_status', 'PASS'))

        ci = 4
        for el in elem_prio:
            rv = pair.get('rpd_values', {}).get(el)
            if rv:
                ws.cell(row=ri, column=ci, value=rv.get('orig'))
                ws.cell(row=ri, column=ci+1, value=rv.get('dup'))
                ws.cell(row=ri, column=ci+2, value=rv.get('rpd'))
                _status_cell(ws, ri, ci+3, rv.get('status', 'PASS'))
            ci += 4

    # Charts: one RPD bar per priority element + regression
    chart_row = len(pairs) + 7
    chart_col_letter = 'A'
    ci = 1
    for el in elem_prio:
        fig = _chart_rpd_bar(pairs, el, thresholds)
        if fig:
            img = _img_to_xl(fig)
            img.width = 600
            img.height = 300
            ws.add_image(img, f'{get_column_letter(ci)}{chart_row}')
            ci += 9

    # Regression charts
    chart_row2 = chart_row + 22
    ci = 1
    for cd in dup_results.get('charts_data', []):
        fig = _chart_scatter_regression(cd)
        if fig:
            img = _img_to_xl(fig)
            img.width = 360
            img.height = 360
            ws.add_image(img, f'{get_column_letter(ci)}{chart_row2}')
            ci += 7

    for ci_w in range(1, len(headers)+1):
        _set_col_width(ws, ci_w, 12)
    _set_col_width(ws, 1, 14)
    _set_col_width(ws, 2, 14)
    ws.freeze_panes = ws.cell(row=5, column=4)


def _write_qc_blanks(wb: Workbook, blk_results: dict, cfg: dict):
    ws = wb.create_sheet('QC_BLANKS')
    ws.sheet_properties.tabColor = TAB_COLORS['QC_BLANKS']
    elem_prio = cfg.get('elements_prioritaires', ['Ta', 'Nb', 'Ti', 'Rb'])

    ws.cell(row=1, column=1, value='QC BLANKS — Contamination Control').font = Font(
        name='Calibri', bold=True, size=14, color=COLORS['cyan'])

    row = 3
    chart_row_start = 3
    for crm_id, crm_res in blk_results.get('crm_results', {}).items():
        ws.cell(row=row, column=1, value=f'CRM: {crm_id}').font = Font(bold=True, size=12)
        ws.cell(row=row, column=3, value=f"Status: {crm_res.get('overall_status')}")
        _status_cell(ws, row, 3, crm_res.get('overall_status', 'PASS'))
        row += 1

        headers = ['ELEMENT', 'N', 'MEAN (ppm)', 'STD', 'CERTIFIED', 'STATUS', 'ISSUES']
        _write_header_row(ws, row, headers, COLORS['cyan'])
        row += 1

        for el, el_res in crm_res.get('elements', {}).items():
            ws.cell(row=row, column=1, value=el)
            ws.cell(row=row, column=2, value=el_res.get('n'))
            ws.cell(row=row, column=3, value=el_res.get('mean'))
            ws.cell(row=row, column=4, value=el_res.get('std'))
            ws.cell(row=row, column=5, value=el_res.get('certified', 'N/A'))
            _status_cell(ws, row, 6, el_res.get('status', 'PASS'))
            ws.cell(row=row, column=7, value=' | '.join(el_res.get('issues', [])))
            row += 1

        row += 2

        # Shewhart charts for priority elements
        chart_col = 9
        for el in elem_prio:
            sd = crm_res.get('shewhart_data', {}).get(el)
            if sd and sd.get('values'):
                fig = _chart_shewhart(sd, crm_id, el, is_recovery=False)
                if fig:
                    img = _img_to_xl(fig)
                    img.width = 600
                    img.height = 240
                    ws.add_image(img, f'{get_column_letter(chart_col)}{chart_row_start}')
                    chart_row_start += 17
                    chart_col = 9

    for ci in range(1, 8):
        _set_col_width(ws, ci, 14)
    _set_col_width(ws, 7, 40)


def _write_qc_standards(wb: Workbook, std_results: dict, cfg: dict):
    ws = wb.create_sheet('QC_STANDARDS')
    ws.sheet_properties.tabColor = TAB_COLORS['QC_STANDARDS']
    elem_prio = cfg.get('elements_prioritaires', ['Ta', 'Nb', 'Ti', 'Rb'])

    is_pxrf = std_results.get('is_pxrf', False)
    pass_range = std_results.get('pass_range', [90, 110])
    warn_range = std_results.get('warn_range', [85, 115])

    ws.cell(row=1, column=1, value='QC STANDARDS — Recovery Analysis').font = Font(
        name='Calibri', bold=True, size=14, color=COLORS['orange'])
    ws.cell(row=2, column=1, value=f"Acceptance: {pass_range[0]}-{pass_range[1]}% | "
            f"Warning: {warn_range[0]}-{warn_range[1]}% | "
            f"{'pXRF (wider tolerance)' if is_pxrf else 'ICP standard tolerance'}")

    row = 4
    chart_row_start = 4
    for crm_id, crm_res in std_results.get('crm_results', {}).items():
        ws.cell(row=row, column=1, value=f'CRM: {crm_id}').font = Font(bold=True, size=12)
        _status_cell(ws, row, 3, crm_res.get('overall_status', 'PASS'))
        row += 1

        if crm_res.get('no_certified'):
            ws.cell(row=row, column=1, value=f"No certified values for: {', '.join(crm_res['no_certified'])}")
            ws.cell(row=row, column=1).font = Font(italic=True, color='888888')
            row += 1

        headers = ['ELEMENT', 'N', 'MEAN_RECOVERY%', 'STD_RECOVERY%',
                   'CERTIFIED (ppm)', 'STATUS', 'DRIFT', 'ISSUES']
        _write_header_row(ws, row, headers, COLORS['orange'])
        row += 1

        for el, el_res in crm_res.get('elements', {}).items():
            ws.cell(row=row, column=1, value=el)
            ws.cell(row=row, column=2, value=el_res.get('n'))
            ws.cell(row=row, column=3, value=el_res.get('mean_recovery'))
            ws.cell(row=row, column=4, value=el_res.get('std_recovery'))
            ws.cell(row=row, column=5, value=el_res.get('certified'))
            _status_cell(ws, row, 6, el_res.get('status', 'PASS'))
            drift = el_res.get('drift', {})
            ws.cell(row=row, column=7, value=drift.get('direction', '') if drift else '')
            ws.cell(row=row, column=8, value=' | '.join(el_res.get('issues', [])))
            row += 1

        row += 2

        # Shewhart recovery charts
        chart_col = 10
        for el in elem_prio:
            sd = crm_res.get('shewhart_data', {}).get(el)
            if sd and sd.get('recoveries'):
                fig = _chart_shewhart(sd, crm_id, el, is_recovery=True)
                if fig:
                    img = _img_to_xl(fig)
                    img.width = 600
                    img.height = 240
                    ws.add_image(img, f'{get_column_letter(chart_col)}{chart_row_start}')
                    chart_row_start += 17

    for ci in range(1, 9):
        _set_col_width(ws, ci, 14)
    _set_col_width(ws, 8, 40)


def _write_historique(wb: Workbook, batches: list):
    ws = wb.create_sheet('HISTORIQUE')
    ws.sheet_properties.tabColor = TAB_COLORS['HISTORIQUE']

    ws.cell(row=1, column=1, value='HISTORIQUE — Batch Archive').font = Font(
        name='Calibri', bold=True, size=14, color=COLORS['gray'])

    if not batches:
        ws.cell(row=3, column=1, value='No previous batches in this zone.')
        return

    headers = ['FILENAME', 'ZONE', 'DATE', 'MODE', 'N_ORIG', 'N_DUP', 'N_BLK', 'N_STD']
    _write_header_row(ws, 3, headers, COLORS['gray'])

    for ri, b in enumerate(batches, start=4):
        ws.cell(row=ri, column=1, value=b.get('filename'))
        ws.cell(row=ri, column=2, value=b.get('zone'))
        ws.cell(row=ri, column=3, value=str(b.get('date', ''))[:19])
        ws.cell(row=ri, column=4, value=b.get('mode'))
        ws.cell(row=ri, column=5, value=b.get('n_orig'))
        ws.cell(row=ri, column=6, value=b.get('n_dup'))
        ws.cell(row=ri, column=7, value=b.get('n_blk'))
        ws.cell(row=ri, column=8, value=b.get('n_std'))

    for ci in range(1, 9):
        _set_col_width(ws, ci, 16)


def _write_qc_lab(wb: Workbook, df_lab: pd.DataFrame, cfg: dict):
    """
    Write QC_LABO sheet — informational only.
    These are QC samples from the lab (QC_Source='LAB'), not from our field program.
    They are displayed for reference but not used for validation.
    """
    ws = wb.create_sheet('QC_LABO')
    ws.sheet_properties.tabColor = TAB_COLORS['gray']
    
    elem_prio = cfg.get('elements_prioritaires', ['Ta', 'Nb', 'Ti', 'Rb'])
    
    # Title
    ws.cell(row=1, column=1, value='QC LABORATOIRE — Informatif').font = Font(
        name='Calibri', bold=True, size=14, color=COLORS['gray'])
    ws.cell(row=2, column=1, value='Ces QC appartiennent au laboratoire et ne sont pas utilisés pour la validation.').font = Font(
        name='Calibri', italic=True, size=9, color=COLORS['muted'])
    
    # Summary stats
    n_lab = len(df_lab)
    n_lab_blk = len(df_lab[df_lab['QAQCType'] == 'BLK']) if 'QAQCType' in df_lab.columns else 0
    n_lab_std = len(df_lab[df_lab['QAQCType'] == 'STD']) if 'QAQCType' in df_lab.columns else 0
    
    ws.cell(row=4, column=1, value=f'Total QC Labo: {n_lab} | Blancs: {n_lab_blk} | Standards: {n_lab_std}').font = Font(bold=True)
    
    # Column selection
    meta_cols = ['SampleID', 'HoleID', 'SiteID', 'ActivityType', 'From_m', 'To_m',
                 'SampleType', 'QAQCType', 'IDBlk', 'IDStd', 'Method', 'Instrument']
    meta_cols = [c for c in meta_cols if c in df_lab.columns]
    
    elem_cols = [f'{e}_ppm' for e in elem_prio if f'{e}_ppm' in df_lab.columns]
    all_cols = meta_cols + elem_cols
    headers = [c.replace('_ppm', '') if c.endswith('_ppm') else c for c in all_cols]
    
    _write_header_row(ws, 6, headers, COLORS['gray'])
    
    # Data rows
    for ri, (_, row) in enumerate(df_lab[all_cols].iterrows(), start=7):
        for ci, col in enumerate(all_cols, 1):
            val = row.get(col)
            if pd.isna(val) or val is None:
                val = ''
            cell = ws.cell(row=ri, column=ci, value=val if not isinstance(val, float) or not np.isnan(val) else '')
            cell.font = _data_font()
            cell.border = _thin_border()
    
    # Column widths
    for ci, col in enumerate(all_cols, 1):
        w = 12 if col.endswith('_ppm') else 14
        _set_col_width(ws, ci, w)
    
    ws.freeze_panes = ws.cell(row=7, column=1)


# ─── MAIN ENTRY POINT ────────────────────────────────────────────────────────

def write_xlsx(df: pd.DataFrame, qc_results: dict, mode: str,
               output_dir: Path, zone: str = None, batches: list = None,
               df_lab: pd.DataFrame = None) -> str:
    """
    Write full XLSX report.
    Returns filepath.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    cfg = load_config()
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    zone_str = zone or 'ZONE'
    filename = f"GeoQC_{zone_str}_Mode{mode}_{now}.xlsx"
    filepath = output_dir / filename

    wb = Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ── Always write DATA_CLEAN ──
    _write_data_clean(wb, df, cfg)

    # ── Conditional sheets ──
    dup_results = qc_results.get('duplicates', {})
    blk_results = qc_results.get('blanks', {})
    std_results = qc_results.get('standards', {})

    if dup_results.get('n_pairs', 0) > 0:
        _write_qc_duplicates(wb, dup_results, cfg)

    if blk_results.get('n_blanks', 0) > 0:
        _write_qc_blanks(wb, blk_results, cfg)

    if std_results.get('n_standards', 0) > 0:
        _write_qc_standards(wb, std_results, cfg)

    # ── QC LABO sheet (informational) ──
    if df_lab is not None and not df_lab.empty:
        _write_qc_lab(wb, df_lab, cfg)

    # ── HISTORIQUE if ≥2 batches ──
    if batches and len(batches) >= 2:
        _write_historique(wb, batches)

    wb.save(filepath)
    return str(filepath)


def load_config() -> dict:
    with open(CONFIG_PATH) as f:
        return json.load(f)
